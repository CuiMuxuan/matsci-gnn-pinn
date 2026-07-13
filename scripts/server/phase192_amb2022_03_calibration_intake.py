#!/usr/bin/env python3
"""Audit checksum-verified AMB2022-03 inputs before any calibration relation is fitted."""

from __future__ import annotations

import argparse
import csv
import json
import os
from pathlib import Path
from typing import Any


DEFAULT_PHASE191 = Path(
    os.environ.get(
        "AMB2022_01_PHASE191_EXTERNAL_CONFIRMATION",
        "/root/matsci-gnn-pinn-ops/phase191_external_confirmation_design.json",
    )
)
DEFAULT_EXTERNAL_ROOT = Path(
    os.environ.get(
        "AMB2022_EXTERNAL_INTAKE_ROOT",
        "/root/matsci-gnn-pinn-data/raw/ambench/2022_external_intake",
    )
)
SIGNAL_RELATIVE_PATH = "AMB2022-03-thermography/Thermography/AMB2022-03-718-AMMT-StaringCamera_Signal.h5"
XYPT_RELATIVE_PATH = "AMB2022-03-thermography/ScanStrategy/AMB2022-03-AMMT-718-Pad_XYPT.h5"
WORKBOOK_RELATIVE_PATH = (
    "AMB2022-03-cross-sections/AMB2022-718-SH1-MeltPool_Cross-Section_Measurement_Results.xlsx"
)
WORKBOOK_REQUIRED_COLUMNS = {
    "Sample",
    "Part No.",
    "Position (mm)",
    "Velocity (mm/s)",
    "Power (W)",
    "Depth (µm)",
    "Width (µm)",
}
INVENTORY_FIELDS = ("source", "identifier", "datasets", "signal_shape", "attributes")


def _h5py() -> Any:
    try:
        import h5py
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError("h5py is required for Phase 192") from exc
    return h5py


def _openpyxl() -> Any:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError("openpyxl is required for Phase 192") from exc
    return load_workbook


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _shape_text(shape: tuple[int, ...]) -> str:
    return "x".join(str(int(value)) for value in shape)


def inspect_thermography(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    h5py = _h5py()
    inventory: list[dict[str, Any]] = []
    with h5py.File(path, "r") as handle:
        top_level = sorted(handle.keys())

        def visitor(name: str, item: Any) -> None:
            if not isinstance(item, h5py.Dataset) or Path(name).name != "Signal":
                return
            group_path = name.rsplit("/", 1)[0] if "/" in name else "/"
            group = handle[group_path]
            inventory.append(
                {
                    "source": "thermography",
                    "identifier": group_path,
                    "datasets": Path(name).name,
                    "signal_shape": _shape_text(item.shape),
                    "attributes": ";".join(sorted(group.attrs.keys())),
                }
            )

        handle.visititems(visitor)
    return {
        "top_level_groups": top_level,
        "signal_group_count": len(inventory),
    }, inventory


def inspect_xypt(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    h5py = _h5py()
    inventory: list[dict[str, Any]] = []
    with h5py.File(path, "r") as handle:
        top_level = sorted(handle.keys())

        def visitor(name: str, item: Any) -> None:
            if not isinstance(item, h5py.Group):
                return
            dataset_names = sorted(key for key, value in item.items() if isinstance(value, h5py.Dataset))
            if not {"X", "Y", "P", "T"}.issubset(dataset_names):
                return
            shapes = {key: _shape_text(item[key].shape) for key in ("X", "Y", "P", "T")}
            inventory.append(
                {
                    "source": "pad_xypt",
                    "identifier": "/" + name if name else "/",
                    "datasets": ";".join(dataset_names),
                    "signal_shape": json.dumps(shapes, sort_keys=True),
                    "attributes": ";".join(sorted(item.attrs.keys())),
                }
            )

        handle.visititems(visitor)
    return {
        "top_level_groups": top_level,
        "xypt_group_count": len(inventory),
    }, inventory


def inspect_workbook(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    load_workbook = _openpyxl()
    workbook = load_workbook(path, read_only=True, data_only=True)
    inventory: list[dict[str, Any]] = []
    all_headers: dict[str, list[str]] = {}
    row_count = 0
    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, ())
        headers = [str(value).strip() for value in header_row if value is not None]
        all_headers[sheet.title] = headers
        data_rows = sum(1 for row in iterator if any(value is not None for value in row))
        row_count += data_rows
        inventory.append(
            {
                "source": "cross_section_workbook",
                "identifier": sheet.title,
                "datasets": ";".join(headers),
                "signal_shape": str(data_rows),
                "attributes": "",
            }
        )
    return {
        "sheet_names": list(all_headers),
        "headers": all_headers,
        "data_row_count": row_count,
    }, inventory


def _phase191_ready(phase191: dict[str, Any]) -> bool:
    gate = phase191.get("gate", {})
    return (
        gate.get("status") == "phase191_external_confirmation_design_ready_phase192_local_calibration_intake"
        and bool(gate.get("phase192_local_calibration_intake_allowed"))
        and gate.get("model_training_allowed") is False
        and gate.get("post_b8_model_reselection_allowed") is False
    )


def build_gate(
    *,
    phase191: dict[str, Any],
    thermography_summary: dict[str, Any],
    xypt_summary: dict[str, Any],
    workbook_summary: dict[str, Any],
) -> dict[str, Any]:
    blockers: list[str] = []
    if not _phase191_ready(phase191):
        blockers.append("phase191_external_confirmation_gate_not_ready")
    if int(thermography_summary.get("signal_group_count", 0)) <= 0:
        blockers.append("thermography_signal_groups_missing")
    if int(xypt_summary.get("xypt_group_count", 0)) <= 0:
        blockers.append("pad_xypt_groups_missing")
    headers = set()
    for values in workbook_summary.get("headers", {}).values():
        headers.update(values)
    missing_headers = sorted(WORKBOOK_REQUIRED_COLUMNS - headers)
    if missing_headers:
        blockers.extend(f"workbook_column_missing_{column}" for column in missing_headers)
    if int(workbook_summary.get("data_row_count", 0)) <= 0:
        blockers.append("cross_section_workbook_rows_missing")
    blockers = sorted(set(blockers))
    ready = not blockers
    return {
        "status": (
            "phase192_amb2022_03_calibration_intake_ready_phase193_identifier_join_design"
            if ready
            else "phase192_amb2022_03_calibration_intake_incomplete"
        ),
        "phase193_identifier_join_design_allowed": ready,
        "calibration_fitting_allowed": False,
        "model_training_allowed": False,
        "hyperparameter_search_allowed": False,
        "post_b8_model_reselection_allowed": False,
        "independent_3d_temperature_confirmation_allowed": False,
        "simulation_as_external_validation_allowed": False,
        "blocking_audits": blockers,
        "next_action": (
            "audit deterministic thermography-XYPT-workbook identifiers before any calibration fit"
            if ready
            else "repair checksum intake or source schema discovery before designing the identifier join"
        ),
    }


def build_intake(
    *,
    phase191: dict[str, Any],
    thermography_summary: dict[str, Any],
    xypt_summary: dict[str, Any],
    workbook_summary: dict[str, Any],
) -> dict[str, Any]:
    return {
        "phase": 192,
        "objective": "checksum_verified_amb2022_03_calibration_schema_intake",
        "thermography": thermography_summary,
        "pad_xypt": xypt_summary,
        "cross_section_workbook": workbook_summary,
        "join_boundary": (
            "No thermography-to-XYPT-to-cross-section row join is assumed from file names or order. "
            "Phase 193 must establish a deterministic identifier mapping."
        ),
        "gate": build_gate(
            phase191=phase191,
            thermography_summary=thermography_summary,
            xypt_summary=xypt_summary,
            workbook_summary=workbook_summary,
        ),
    }


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=INVENTORY_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--phase191", type=Path, default=DEFAULT_PHASE191)
    parser.add_argument("--external-root", type=Path, default=DEFAULT_EXTERNAL_ROOT)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--inventory-csv", type=Path, required=True)
    args = parser.parse_args()
    thermography_summary, thermography_rows = inspect_thermography(args.external_root / SIGNAL_RELATIVE_PATH)
    xypt_summary, xypt_rows = inspect_xypt(args.external_root / XYPT_RELATIVE_PATH)
    workbook_summary, workbook_rows = inspect_workbook(args.external_root / WORKBOOK_RELATIVE_PATH)
    payload = build_intake(
        phase191=_read_json(args.phase191),
        thermography_summary=thermography_summary,
        xypt_summary=xypt_summary,
        workbook_summary=workbook_summary,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    _write_csv(args.inventory_csv, thermography_rows + xypt_rows + workbook_rows)
    print(json.dumps(payload["gate"], indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
