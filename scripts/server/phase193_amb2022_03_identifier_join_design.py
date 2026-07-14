#!/usr/bin/env python3
"""Build an auditable AMB2022-03 single-track condition join without fitting a model."""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import numpy as np


DEFAULT_PHASE192 = Path(
    os.environ.get(
        "AMB2022_01_PHASE192_CALIBRATION_INTAKE",
        "/root/matsci-gnn-pinn-ops/phase192_amb2022_03_calibration_intake.json",
    )
)
DEFAULT_EXTERNAL_ROOT = Path(
    os.environ.get(
        "AMB2022_EXTERNAL_INTAKE_ROOT",
        "/root/matsci-gnn-pinn-data/raw/ambench/2022_external_intake",
    )
)
SIGNAL_RELATIVE_PATH = "AMB2022-03-thermography/Thermography/AMB2022-03-718-AMMT-StaringCamera_Signal.h5"
WORKBOOK_RELATIVE_PATH = (
    "AMB2022-03-cross-sections/AMB2022-718-SH1-MeltPool_Cross-Section_Measurement_Results.xlsx"
)
CONDITION_FIELDS = (
    "thermal_group_id",
    "workbook_case_id",
    "workbook_sample_id",
    "n_cross_section_replicates",
    "laser_power_W",
    "scan_speed_mm_s",
    "spot_size_um",
    "depth_um_mean",
    "depth_um_std",
    "width_um_mean",
    "width_um_std",
)
PROCESS_COLUMNS = {
    "laser_power_W": "Power (W)",
    "scan_speed_mm_s": "Velocity (mm/s)",
    "spot_size_um": "Beam diameter (gauss, avg) (µm)",
}


def _h5py() -> Any:
    try:
        import h5py
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError("h5py is required for Phase 193") from exc
    return h5py


def _openpyxl() -> Any:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError("openpyxl is required for Phase 193") from exc
    return load_workbook


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def normalize_condition_id(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def _scalar_attribute(group: Any, key: str) -> float:
    values = np.asarray(group.attrs[key]).reshape(-1)
    if values.size != 1:
        raise ValueError(f"Expected scalar {key} in {group.name}, found shape {values.shape}")
    return float(values[0])


def _mean_std(values: list[float]) -> tuple[float, float]:
    mean = sum(values) / len(values)
    if len(values) < 2:
        return mean, 0.0
    return mean, math.sqrt(sum((value - mean) ** 2 for value in values) / (len(values) - 1))


def _thermal_conditions(path: Path) -> tuple[dict[str, dict[str, float]], list[str]]:
    h5py = _h5py()
    conditions: dict[str, dict[str, float]] = {}
    excluded_pad_ids: list[str] = []
    with h5py.File(path, "r") as handle:
        def visitor(name: str, item: Any) -> None:
            if not isinstance(item, h5py.Dataset) or Path(name).name != "Signal":
                return
            group = handle[name.rsplit("/", 1)[0]]
            identifier = group.name.rsplit("/", 1)[-1]
            if identifier.startswith("Line_"):
                if identifier in conditions:
                    raise ValueError(f"Duplicate thermal condition: {identifier}")
                conditions[identifier] = {
                    "laser_power_W": _scalar_attribute(group, "laser_power"),
                    "scan_speed_mm_s": _scalar_attribute(group, "scan_speed"),
                    "spot_size_um": _scalar_attribute(group, "spot_size"),
                }
            else:
                excluded_pad_ids.append(identifier)

        handle.visititems(visitor)
    return conditions, sorted(excluded_pad_ids)


def _workbook_rows(path: Path) -> list[dict[str, Any]]:
    load_workbook = _openpyxl()
    sheet = load_workbook(path, read_only=True, data_only=True).active
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() for value in next(iterator, ())]
    return [dict(zip(headers, row)) for row in iterator if any(value is not None for value in row)]


def build_condition_join(
    thermal_conditions: dict[str, dict[str, float]], workbook_rows: list[dict[str, Any]], excluded_pad_ids: list[str]
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    single_track_rows = [row for row in workbook_rows if str(row.get("Sample", "")).endswith("BP1")]
    by_case: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in single_track_rows:
        by_case[normalize_condition_id(str(row.get("Case and Line No.", "")))].append(row)
    joined: list[dict[str, Any]] = []
    unmatched_thermal: list[str] = []
    process_mismatches: list[str] = []
    used_cases: set[str] = set()
    for thermal_id, process in sorted(thermal_conditions.items()):
        normalized = normalize_condition_id(thermal_id)
        candidates = by_case.get(normalized, [])
        if not candidates:
            unmatched_thermal.append(thermal_id)
            continue
        compatible = True
        for row in candidates:
            for output_name, source_name in PROCESS_COLUMNS.items():
                if not math.isclose(float(row[source_name]), process[output_name], rel_tol=0.0, abs_tol=1e-9):
                    compatible = False
        if not compatible:
            process_mismatches.append(thermal_id)
            continue
        depth_values = [float(row["Depth (µm)"]) for row in candidates]
        width_values = [float(row["Width (µm)"]) for row in candidates]
        depth_mean, depth_std = _mean_std(depth_values)
        width_mean, width_std = _mean_std(width_values)
        first = candidates[0]
        joined.append(
            {
                "thermal_group_id": thermal_id,
                "workbook_case_id": str(first["Case and Line No."]),
                "workbook_sample_id": str(first["Sample"]),
                "n_cross_section_replicates": len(candidates),
                **process,
                "depth_um_mean": depth_mean,
                "depth_um_std": depth_std,
                "width_um_mean": width_mean,
                "width_um_std": width_std,
            }
        )
        used_cases.add(normalized)
    extra_workbook_cases = sorted(set(by_case) - used_cases)
    audit = {
        "thermal_single_track_count": len(thermal_conditions),
        "workbook_single_track_row_count": len(single_track_rows),
        "joined_condition_count": len(joined),
        "unmatched_thermal_ids": unmatched_thermal,
        "process_mismatch_thermal_ids": process_mismatches,
        "extra_workbook_case_keys": extra_workbook_cases,
        "excluded_pad_signal_ids": sorted(excluded_pad_ids),
    }
    return joined, audit


def _phase192_ready(phase192: dict[str, Any]) -> bool:
    gate = phase192.get("gate", {})
    return (
        gate.get("status") == "phase192_amb2022_03_calibration_intake_ready_phase193_identifier_join_design"
        and bool(gate.get("phase193_identifier_join_design_allowed"))
        and gate.get("calibration_fitting_allowed") is False
        and gate.get("model_training_allowed") is False
    )


def build_gate(phase192: dict[str, Any], audit: dict[str, Any], join_rows: list[dict[str, Any]]) -> dict[str, Any]:
    blockers: list[str] = []
    if not _phase192_ready(phase192):
        blockers.append("phase192_calibration_intake_gate_not_ready")
    if int(audit.get("thermal_single_track_count", 0)) != int(audit.get("joined_condition_count", 0)):
        blockers.append("thermal_condition_join_incomplete")
    if audit.get("unmatched_thermal_ids"):
        blockers.append("thermal_identifier_unmatched")
    if audit.get("process_mismatch_thermal_ids"):
        blockers.append("thermal_workbook_process_mismatch")
    if audit.get("extra_workbook_case_keys"):
        blockers.append("workbook_single_track_case_unmatched")
    if not join_rows or any(int(row["n_cross_section_replicates"]) != 2 for row in join_rows):
        blockers.append("cross_section_replicate_contract_not_two")
    blockers = sorted(set(blockers))
    ready = not blockers
    return {
        "status": (
            "phase193_identifier_join_design_ready_phase194_calibration_protocol_design"
            if ready
            else "phase193_identifier_join_design_incomplete_or_ambiguous"
        ),
        "phase194_calibration_protocol_design_allowed": ready,
        "calibration_fitting_allowed": False,
        "model_training_allowed": False,
        "hyperparameter_search_allowed": False,
        "post_b8_model_reselection_allowed": False,
        "excluded_pad_signal_count": len(audit.get("excluded_pad_signal_ids", [])),
        "independent_3d_temperature_confirmation_allowed": False,
        "simulation_as_external_validation_allowed": False,
        "blocking_audits": blockers,
        "next_action": (
            "freeze a calibration-versus-stress-test protocol for the 21 joined single-track conditions before fitting"
            if ready
            else "repair ambiguous identifiers or process metadata before constructing any calibration target"
        ),
    }


def build_design(
    *, phase192: dict[str, Any], join_rows: list[dict[str, Any]], audit: dict[str, Any]
) -> dict[str, Any]:
    return {
        "phase": 193,
        "objective": "deterministic_amb2022_03_single_track_identifier_join",
        "condition_join": join_rows,
        "join_audit": audit,
        "join_boundary": (
            "Only single-track Line_* groups with exact normalized workbook identifiers and matching process metadata are joined. "
            "Pad groups remain excluded until a separate identifier audit succeeds."
        ),
        "gate": build_gate(phase192, audit, join_rows),
    }


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CONDITION_FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--phase192", type=Path, default=DEFAULT_PHASE192)
    parser.add_argument("--external-root", type=Path, default=DEFAULT_EXTERNAL_ROOT)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--conditions-csv", type=Path, required=True)
    args = parser.parse_args()
    thermal_conditions, excluded_pad_ids = _thermal_conditions(args.external_root / SIGNAL_RELATIVE_PATH)
    workbook_rows = _workbook_rows(args.external_root / WORKBOOK_RELATIVE_PATH)
    join_rows, audit = build_condition_join(thermal_conditions, workbook_rows, excluded_pad_ids)
    payload = build_design(phase192=_read_json(args.phase192), join_rows=join_rows, audit=audit)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    _write_csv(args.conditions_csv, join_rows)
    print(json.dumps(payload["gate"], indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
